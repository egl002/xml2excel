import io
import xml.etree.ElementTree as ET
from datetime import datetime, time
from typing import Optional, Dict, Tuple, List

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import numbers


st.set_page_config(page_title="XML2Excel", layout="centered")

st.title("XML2Excel")
st.caption("Last opp XML-startliste fra EQ Timing. Få tilbake KES-vennlig Excel. By Espen at Geilo IL.")

uploaded = st.file_uploader("Last opp XML-fil", type=["xml"])

mode = st.radio(
    "Klasse → tall",
    options=["Først-seen i XML (anbefalt)", "Alfabetisk"],
    index=0
)


def clean_ws(s: str) -> str:
    return " ".join((s or "").split())


def parse_time_maybe(s: str) -> Optional[time]:
    if not s:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    return None


def parse_xml(xml_bytes: bytes) -> pd.DataFrame:
    rows: List[List[str]] = []
    f = io.BytesIO(xml_bytes)

    for _, elem in ET.iterparse(f, events=("end",)):
        if elem.tag != "start":
            continue

        a = elem.attrib
        startno = a.get("startno", "")
        fornavn = clean_ws(a.get("fornavn", ""))
        etternavn = clean_ws(a.get("etternavn", ""))
        navn = f"{fornavn} {etternavn}".strip()

        klasse = a.get("klasse", "")
        team = clean_ws(a.get("team", ""))
        starttid = a.get("starttid", "")

        rows.append([startno, navn, klasse, team, starttid])
        elem.clear()

    return pd.DataFrame(rows, columns=["startno", "navn", "klasse", "team", "starttid"])


def build_class_ids(df_raw: pd.DataFrame, class_mode: str) -> Tuple[pd.DataFrame, Dict[str, int]]:
    df = df_raw.copy()
    df["klasse"] = df["klasse"].fillna("").astype(str)

    if class_mode == "Alfabetisk":
        unique_classes = sorted(df["klasse"].unique())
        mapping = {c: i for i, c in enumerate(unique_classes)}
    else:
        mapping = {}
        next_id = 0
        for c in df["klasse"]:
            if c not in mapping:
                mapping[c] = next_id
                next_id += 1

    df["klasse_id"] = df["klasse"].map(mapping)

    out = df[["startno", "navn", "klasse_id", "team", "starttid"]].copy()

    out["startno"] = pd.to_numeric(out["startno"], errors="coerce").astype("Int64")
    out["klasse_id"] = pd.to_numeric(out["klasse_id"], errors="coerce").astype("Int64")
    out["starttid"] = out["starttid"].apply(parse_time_maybe)

    return out, mapping


def build_excel(df_out: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active

    for row in df_out.itertuples(index=False, name=None):
        ws.append(list(row))

    # Tving tallformat
    for r in range(1, ws.max_row + 1):
        if isinstance(ws.cell(r, 1).value, int):
            ws.cell(r, 1).number_format = numbers.FORMAT_NUMBER
        if isinstance(ws.cell(r, 3).value, int):
            ws.cell(r, 3).number_format = numbers.FORMAT_NUMBER

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


if uploaded:
    df_raw = parse_xml(uploaded.read())

    if df_raw.empty:
        st.error("Fant ingen startposter.")
        st.stop()

    df_out, mapping = build_class_ids(df_raw, mode)

    st.success(f"Fant {len(df_out)} starter og {len(mapping)} klasser.")
    st.dataframe(df_out.head(20), use_container_width=True)

    xlsx = build_excel(df_out)

    st.download_button(
        "Last ned Excel (.xlsx)",
        xlsx,
        "startliste_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
st.divider()
st.subheader("Manualer")

st.markdown(
    """
    📘 [Last ned KES-manual (PDF)](https://jottacloud.com/s/395dd442cd79c9744f89a95857f25f0a165)

    📘 [Last ned Keyboard Manual (PDF)](https://jottacloud.com/s/3956406dbaf42f045769a18aff6a71db44f)

    📗 [Last ned KES opppsett --> EQTiming (PDF)](https://jottacloud.com/s/395a3e1e1eeb52e4e3d8756cde717634a82)
    """
)
